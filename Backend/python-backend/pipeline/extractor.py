import csv
import io
import os
import tempfile
import requests
import fitz
from docx import Document as DocxDocument

try:
    from docling.document_converter import DocumentConverter
except ImportError:
    DocumentConverter = None


_DOC_CONVERTER = None


def detect_document_type(name: str, content_type: str = "") -> str:
    return _detect_extension(name, content_type)


def extract_text(url: str, doc_name: str) -> str:
    response = requests.get(url, timeout=120)
    response.raise_for_status()
    content = response.content

    ext = _detect_extension(doc_name, response.headers.get("content-type", ""))

    if ext == "pdf":
        return _extract_markdown_document(content, doc_name, ext)
    elif ext == "docx":
        return _extract_markdown_document(content, doc_name, ext)
    elif ext in {"csv", "xlsx", "xls"}:
        return _extract_tabular(content, ext)
    elif ext == "txt":
        return content.decode("utf-8", errors="replace")
    else:
        return _extract_pdf(content)


def _detect_extension(name: str, content_type: str) -> str:
    name_lower = name.lower()
    if name_lower.endswith(".pdf"):
        return "pdf"
    elif name_lower.endswith(".docx"):
        return "docx"
    elif name_lower.endswith(".csv"):
        return "csv"
    elif name_lower.endswith(".xlsx"):
        return "xlsx"
    elif name_lower.endswith(".xls"):
        return "xls"
    elif name_lower.endswith(".txt"):
        return "txt"

    ct = content_type.lower()
    if "pdf" in ct:
        return "pdf"
    elif "wordprocessingml" in ct or "msword" in ct:
        return "docx"
    elif "csv" in ct:
        return "csv"
    elif "excel" in ct or "spreadsheet" in ct or "sheet" in ct:
        return "xlsx"
    elif "text/plain" in ct:
        return "txt"

    return "pdf"


def _extract_pdf(content: bytes) -> str:
    doc = fitz.open(stream=content, filetype="pdf")
    pages = []
    for page in doc:
        text = page.get_text("text")
        if text.strip():
            pages.append(text)
    doc.close()
    return "\n\n".join(pages)


def _extract_docx(content: bytes) -> str:
    buf = io.BytesIO(content)
    doc = DocxDocument(buf)
    paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]
    return "\n\n".join(paragraphs)


def _extract_markdown_document(content: bytes, doc_name: str, ext: str) -> str:
    markdown_text = _extract_with_docling_converter(content, doc_name, ext)
    if markdown_text:
        return markdown_text

    if ext == "pdf":
        return _extract_pdf(content)

    return _extract_docx(content)


def _extract_with_docling_converter(content: bytes, doc_name: str, ext: str) -> str:
    if DocumentConverter is None:
        return ""

    suffix = f".{ext}" if ext else ""
    temp_path = None

    try:
        with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as temp_file:
            temp_file.write(content)
            temp_path = temp_file.name

        converter = _get_docling_converter()
        if converter is None:
            return ""

        result = converter.convert(temp_path)
        document = getattr(result, "document", None)
        if document is None:
            return ""

        markdown_text = document.export_to_markdown()
        if isinstance(markdown_text, str):
            return markdown_text.strip()

        return str(markdown_text).strip()
    except Exception:
        return ""
    finally:
        if temp_path and os.path.exists(temp_path):
            try:
                os.unlink(temp_path)
            except OSError:
                pass


def _get_docling_converter():
    global _DOC_CONVERTER

    if _DOC_CONVERTER is None and DocumentConverter is not None:
        _DOC_CONVERTER = DocumentConverter()

    return _DOC_CONVERTER


def _extract_tabular(content: bytes, ext: str) -> str:
    rows = _load_tabular_rows(content, ext)
    if not rows:
        return ""

    headers = [str(cell).strip() if cell is not None else "" for cell in rows[0]]
    if not any(headers):
        headers = [f"Column {index + 1}" for index in range(len(rows[0]))]
        data_rows = rows
        start_row_number = 1
    else:
        data_rows = rows[1:]
        start_row_number = 2

    formatted_rows = []
    for row_index, row in enumerate(data_rows, start=start_row_number):
        formatted_rows.append(_format_tabular_row(row_index, headers, row))

    return "\n".join(formatted_rows)


def _load_tabular_rows(content: bytes, ext: str) -> list:
    if ext == "csv":
        text = content.decode("utf-8-sig", errors="replace")
        reader = csv.reader(io.StringIO(text))
        return [row for row in reader if any(str(cell).strip() for cell in row)]

    if ext in {"xlsx", "xls"}:
        if ext == "xls":
            raise ValueError("XLS files are not supported yet. Convert the file to XLSX or CSV.")

        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise ImportError("openpyxl is required to read XLSX files.") from exc

        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        worksheet = workbook.active
        rows = []
        for row in worksheet.iter_rows(values_only=True):
            values = [cell for cell in row]
            if any(str(cell).strip() for cell in values if cell is not None):
                rows.append(values)
        workbook.close()
        return rows

    return []


def _format_tabular_row(row_number: int, headers: list, row: list) -> str:
    max_columns = max(len(headers), len(row))
    entries = []

    for index in range(max_columns):
        header = headers[index] if index < len(headers) else f"Column {index + 1}"
        value = row[index] if index < len(row) else ""
        value_text = "" if value is None else str(value).strip()
        entries.append(f"{header}: {value_text}")

    end_column = _excel_column_name(max_columns)
    row_label = f"A{row_number}:{end_column}{row_number}"
    return f"Row {row_number} ({row_label}) | " + " | ".join(entries)


def _excel_column_name(index: int) -> str:
    if index < 1:
        return "A"

    name = ""
    while index > 0:
        index, remainder = divmod(index - 1, 26)
        name = chr(65 + remainder) + name
    return name
